import calendar
import os
import sqlite3
from datetime import datetime, timedelta
from functools import wraps

from flask import Flask, render_template, request, redirect, url_for, flash, send_file
from flask_login import LoginManager, UserMixin, login_user, login_required, logout_user, current_user
from openpyxl import Workbook
from openpyxl.styles import Font
from werkzeug.security import check_password_hash, generate_password_hash
from werkzeug.utils import secure_filename

BASE = os.path.dirname(__file__)
DB   = os.path.join(BASE, "database", "talma_tea.db")
UPLOADS = os.path.join(BASE, "uploads", "vouchers")
EXPORTS = os.path.join(BASE, "exports")
os.makedirs(UPLOADS, exist_ok=True)
os.makedirs(EXPORTS, exist_ok=True)

app = Flask(__name__)
app.secret_key = "talma-prime-2025"

login_manager = LoginManager(app)
login_manager.login_view = "login"

# ----------------------- Helpers -----------------------

def db():
    conn = sqlite3.connect(DB)
    conn.row_factory = sqlite3.Row
    return conn

def _ym(): return datetime.now().strftime("%Y-%m")
def _today(): return datetime.now().strftime("%Y-%m-%d")
def _f(v):
    try: return float(v or 0)
    except: return 0.0

def roles_required(*roles):
    def deco(fn):
        @wraps(fn)
        def inner(*a, **kw):
            if not current_user.is_authenticated:
                return login_manager.unauthorized()
            if current_user.role not in roles:
                flash("Not allowed.", "danger")
                return redirect(url_for("welcome"))
            return fn(*a, **kw)
        return inner
    return deco

class User(UserMixin):
    def __init__(self, r):
        self.id = r["id"]
        self.username = r["username"]
        self.role = r["role"]

@login_manager.user_loader
def load_user(user_id):
    c = db(); r = c.execute("SELECT * FROM users WHERE id=?", (user_id,)).fetchone(); c.close()
    return User(r) if r else None

# ----------------------- Auth / Home -----------------------

@app.route("/login", methods=["GET","POST"])
def login():
    if request.method == "POST":
        u = request.form.get("username","").strip()
        p = request.form.get("password","")
        c = db(); r = c.execute("SELECT * FROM users WHERE username=?", (u,)).fetchone(); c.close()
        if r and check_password_hash(r["password"], p):
            login_user(User(r))
            return redirect(url_for("welcome"))
        flash("Invalid credentials", "danger")
    return render_template("login.html")

@app.route("/logout")
@login_required
def logout():
    logout_user(); return redirect(url_for("login"))

@app.route("/")
@login_required
def welcome():
    return render_template("welcome.html")

# ----------------------- Daily Inputs -----------------------

@app.route("/daily", methods=["GET", "POST"])
@login_required
def daily():
    """Daily inputs grid (month view) + create a new row."""
    month = request.args.get("month") or _ym()  # YYYY-MM

    if request.method == "POST":
        g = request.form.get
        date_str = g("date") or _today()

        # numeric fields
        total_kg   = _f(g("total_kg"))
        amount_tk  = _f(g("amount_tk"))
        gl_pay     = _f(g("green_leaf_bill_payment"))
        staff_salary   = _f(g("staff_salary"))
        labour_bill    = _f(g("labour_bill"))
        production_cost= _f(g("production_cost"))
        coal           = _f(g("coal"))
        diesel         = _f(g("diesel"))
        electricity    = _f(g("electricity"))
        other_exp      = _f(g("other_exp"))
        capital_cost   = _f(g("capital_cost"))
        machineries    = _f(g("machineries"))
        assets_purchase= _f(g("assets_purchase"))
        construction   = _f(g("construction"))
        cash_receive   = _f(g("cash_receive"))
        add_amount     = _f(g("add_amount"))
        less_amount    = _f(g("less_amount"))

        # derived totals
        total_cost = gl_pay + staff_salary + labour_bill + production_cost + coal + diesel + electricity + other_exp
        fixed_cost = capital_cost + machineries + assets_purchase + construction
        grand_total = total_cost + fixed_cost
        net_cash = cash_receive + add_amount - less_amount

        note = (g("note") or "").strip()

        # optional voucher upload
        vp = None
        file = request.files.get("voucher")
        if file and file.filename:
            safe = secure_filename(file.filename)
            fname = f"{datetime.now().strftime('%Y%m%d_%H%M%S')}_{safe}"
            file.save(os.path.join(UPLOADS, fname))
            vp = fname

        # INSERT (named params)
        conn = db()
        vals = dict(
            date=date_str,
            total_kg=total_kg, amount_tk=amount_tk, green_leaf_bill_payment=gl_pay,
            staff_salary=staff_salary, labour_bill=labour_bill, production_cost=production_cost,
            coal=coal, diesel=diesel, electricity=electricity, other_exp=other_exp, total_cost=total_cost,
            capital_cost=capital_cost, machineries=machineries, assets_purchase=assets_purchase,
            construction=construction, fixed_cost=fixed_cost, grand_total=grand_total,
            cash_receive=cash_receive, add_amount=add_amount, less_amount=less_amount, net_cash=net_cash,
            note=note, voucher_path=vp, status="submitted", created_by=current_user.id
        )
        conn.execute("""
          INSERT INTO daily_cash(
            date, total_kg, amount_tk, green_leaf_bill_payment,
            staff_salary, labour_bill, production_cost, coal, diesel, electricity, other_exp, total_cost,
            capital_cost, machineries, assets_purchase, construction, fixed_cost, grand_total,
            cash_receive, add_amount, less_amount, net_cash,
            note, voucher_path, status, created_by
          ) VALUES (
            :date, :total_kg, :amount_tk, :green_leaf_bill_payment,
            :staff_salary, :labour_bill, :production_cost, :coal, :diesel, :electricity, :other_exp, :total_cost,
            :capital_cost, :machineries, :assets_purchase, :construction, :fixed_cost, :grand_total,
            :cash_receive, :add_amount, :less_amount, :net_cash,
            :note, :voucher_path, :status, :created_by
          )
        """, vals)
        conn.commit(); conn.close()

        flash("Saved (awaiting MD/Admin approval).", "success")
        return redirect(url_for("daily", month=date_str[:7]))

    # GET: rows + totals for month
    conn = db()
    rows = conn.execute("""
        SELECT * FROM daily_cash
        WHERE substr(date,1,7)=?
        ORDER BY date ASC, id ASC
    """, (month,)).fetchall()

    tot = conn.execute("""
        SELECT
          COALESCE(SUM(total_kg),0)                    AS total_kg,
          COALESCE(SUM(amount_tk),0)                   AS amount_tk,
          COALESCE(SUM(green_leaf_bill_payment),0)     AS green_leaf_bill_payment,
          COALESCE(SUM(staff_salary),0)                AS staff_salary,
          COALESCE(SUM(labour_bill),0)                 AS labour_bill,
          COALESCE(SUM(production_cost),0)             AS production_cost,
          COALESCE(SUM(coal),0)                        AS coal,
          COALESCE(SUM(diesel),0)                      AS diesel,
          COALESCE(SUM(electricity),0)                 AS electricity,
          COALESCE(SUM(other_exp),0)                   AS other_exp,
          COALESCE(SUM(total_cost),0)                  AS total_cost,
          COALESCE(SUM(capital_cost),0)                AS capital_cost,
          COALESCE(SUM(machineries),0)                 AS machineries,
          COALESCE(SUM(assets_purchase),0)             AS assets_purchase,
          COALESCE(SUM(construction),0)                AS construction,
          COALESCE(SUM(fixed_cost),0)                  AS fixed_cost,
          COALESCE(SUM(grand_total),0)                 AS grand_total,
          COALESCE(SUM(cash_receive),0)                AS cash_receive,
          COALESCE(SUM(add_amount),0)                  AS add_amount,
          COALESCE(SUM(less_amount),0)                 AS less_amount,
          COALESCE(SUM(net_cash),0)                    AS net_cash
        FROM daily_cash
        WHERE substr(date,1,7)=?
    """, (month,)).fetchone()
    conn.close()

    return render_template("daily.html", month=month, rows=rows, tot=tot, today=_today())

@app.route("/daily/approve/<int:row_id>")
@login_required
@roles_required('admin','md')
def approve_daily(row_id):
    conn = db()
    conn.execute("UPDATE daily_cash SET status='approved', approved_by=?, approved_at=? WHERE id=?",
                 (current_user.id, datetime.now().isoformat(timespec='seconds'), row_id))
    conn.commit(); conn.close()
    flash("Daily entry approved.", "success")
    return redirect(request.referrer or url_for("daily"))

@app.route("/daily/unapprove/<int:row_id>")
@login_required
@roles_required('admin','md')
def unapprove_daily(row_id):
    conn = db()
    conn.execute("UPDATE daily_cash SET status='submitted', approved_by=NULL, approved_at=NULL WHERE id=?", (row_id,))
    conn.commit(); conn.close()
    flash("Daily entry reset to submitted.", "info")
    return redirect(request.referrer or url_for("daily"))

# Reject = DELETE (remove from system)
@app.route("/daily/delete/<int:row_id>")
@login_required
@roles_required('admin','md')
def delete_daily(row_id):
    conn = db()
    conn.execute("DELETE FROM daily_cash WHERE id=?", (row_id,))
    conn.commit(); conn.close()
    flash("Daily entry deleted.", "warning")
    return redirect(request.referrer or url_for("daily"))

# Back-compat: if any template still calls /daily/reject/<id>, treat as delete.
@app.route("/daily/reject/<int:row_id>")
@login_required
@roles_required('admin','md')
def reject_daily(row_id):
    return delete_daily(row_id)

# ----------------------- People (Workers & Staff) -----------------------

# --- Add Worker Route --- #
@app.route("/people", methods=["GET", "POST"])
@login_required
@roles_required('manager', 'md', 'admin')
def people():
    conn = db()

    if request.method == "POST":
        kind = request.form.get("kind")
        if kind == "worker":
            # Insert worker with hourly_rate
            conn.execute("INSERT INTO workers(name, join_date, note, active, hourly_rate, approved_hourly_rate) VALUES(?,?,?,?,?,?)",
                         (request.form["name"], request.form.get("join_date") or _today(), request.form.get("note", ""), 1, request.form.get("hourly_rate", "0"), 0))
            conn.commit()
            flash("Worker added.", "success")
        elif kind == "staff":
            # Insert staff with salary
            conn.execute("INSERT INTO staff(name, position, salary, join_date, approved_salary) VALUES(?,?,?,?,?)",
                         (request.form["name"], request.form.get("position", ""), _f(request.form.get("salary")), request.form.get("join_date") or _today(), None))
            conn.commit()
            flash("Staff added.", "success")

        return redirect(url_for("people"))

    workers = conn.execute("SELECT * FROM workers ORDER BY active DESC, name").fetchall()
    staff = conn.execute("SELECT * FROM staff ORDER BY name").fetchall()

    workers = conn.execute("SELECT * FROM workers ORDER BY active DESC, name").fetchall()

    # --- Calculate Weekly Wages for Workers --- #
    # For this, we assume each worker works 40 hours per week as an example
    worker_list = []
    for w in workers:
        # Convert sqlite3.Row to a regular dictionary (if needed)
        worker_data = dict(w)  # Convert the Row to a dictionary
        hourly_rate = float(worker_data.get('hourly_rate', 0))  # Ensure it's a float
        worker_data['weekly_wages'] = hourly_rate * 40  # Assuming 40 hours per week
        worker_list.append(worker_data)

    # Now `worker_list` will contain dictionaries with all the worker data, including weekly_wages

    # --- Calculate Salaries for Staff --- #
    # Use dictionary-like access to get the salary
    total_staff_salary = sum(s['salary'] for s in staff)

    conn.close()
    return render_template("people.html", workers=workers, staff=staff, total_staff_salary=total_staff_salary)


@app.route("/people/update_hourly_rate/<int:worker_id>", methods=["POST"])
@login_required
@roles_required('md', 'admin')
def update_hourly_rate(worker_id):
    new_rate = request.form.get("new_rate")
    conn = db()

    # Check if the user is allowed to approve the rate change
    if current_user.role in ['md', 'admin']:
        # Update the worker's hourly rate but mark it as pending approval
        conn.execute("UPDATE workers SET hourly_rate = ?, approved_hourly_rate = 0 WHERE id = ?",
                     (new_rate, worker_id))
        conn.commit()
        flash("Hourly rate updated (pending approval).", "success")

    conn.close()
    return redirect(url_for("people"))


@app.route("/people/approve_hourly_rate/<int:worker_id>")
@login_required
@roles_required('md', 'admin')
def approve_hourly_rate(worker_id):
    conn = db()

    # Approve the hourly rate change and set it to approved_hourly_rate
    conn.execute("UPDATE workers SET approved_hourly_rate = hourly_rate WHERE id = ?", (worker_id,))
    conn.commit()
    flash("Hourly rate approved.", "success")
    conn.close()

    return redirect(url_for("people"))


@app.route("/people/update_salary/<int:staff_id>", methods=["POST"])
@login_required
@roles_required('md', 'admin')
def update_salary(staff_id):
    new_salary = request.form.get("new_salary")
    conn = db()

    # Update the salary but leave it pending approval
    conn.execute("UPDATE staff SET salary = ?, approved_salary = NULL WHERE id = ?",
                 (new_salary, staff_id))
    conn.commit()
    flash("Salary updated (pending approval).", "success")
    conn.close()
    return redirect(url_for("people"))


@app.route("/people/approve_salary/<int:staff_id>")
@login_required
@roles_required('md', 'admin')
def approve_salary(staff_id):
    conn = db()

    # Approve the salary change
    conn.execute("UPDATE staff SET approved_salary = salary WHERE id = ?", (staff_id,))
    conn.commit()
    flash("Salary approved.", "success")
    conn.close()

    return redirect(url_for("people"))


@app.route("/people/mark_left/<int:worker_id>")
@login_required
@roles_required('md','admin')
def mark_left(worker_id):
    conn = db()
    conn.execute("UPDATE workers SET active=0, leave_date=? WHERE id=?", (_today(), worker_id))
    conn.commit(); conn.close()
    flash("Worker marked as left.", "info")
    return redirect(url_for("people"))

@app.route("/people/delete/<kind>/<int:pid>")
@login_required
@roles_required('md','admin')
def delete_person(kind, pid):
    conn = db()
    if kind=="worker": conn.execute("DELETE FROM workers WHERE id=?", (pid,))
    else:              conn.execute("DELETE FROM staff WHERE id=?", (pid,))
    conn.commit(); conn.close()
    flash("Deleted.", "warning")
    return redirect(url_for("people"))

# ----------------------- Timesheets (Working Hours) -----------------------
def get_week_start(date_str):
    date_obj = datetime.strptime(date_str, "%Y-%m-%d")
    weekday = date_obj.weekday()
    days_to_saturday = (5 - weekday) % 7  # Saturday is day 5 in Python's weekday()
    saturday = date_obj + timedelta(days=days_to_saturday)
    return saturday.strftime("%Y-%m-%d")  # Return the Saturday as the start of the week

@app.route("/timesheets", methods=["GET","POST"])
@login_required
@roles_required('manager','md','admin')
def timesheets_page():
    the_day = request.args.get("date") or _today()
    month = the_day[:7]  # YYYY-MM
    conn = db()

    # Workers for entry form (active)
    workers = conn.execute(
        "SELECT id, name FROM workers WHERE active=1 ORDER BY name"
    ).fetchall()

    # Save entries for the chosen day
    if request.method == "POST":
        for w in workers:
            key = f"hours_{w['id']}"
            h = _f(request.form.get(key))
            if h > 0:
                conn.execute("""
                  INSERT INTO timesheets(date, worker_id, hours, note, status, created_by)
                  VALUES(?,?,?,?, 'pending', ?)
                """, (the_day, w["id"], h, request.form.get(f"note_{w['id']}",""), current_user.id))
        conn.commit()
        flash("Saved timesheets (pending approval).", "success")
        conn.close()
        return redirect(url_for("timesheets_page", date=the_day))

    # Submitted rows for the visible day
    rows = conn.execute("""
      SELECT t.*, w.name
      FROM timesheets t
      JOIN workers w ON w.id=t.worker_id
      WHERE t.date=?
      ORDER BY w.name
    """, (the_day,)).fetchall()

    # Weekly sums per worker/week (group by week start date)
    week_rows = conn.execute("""
        SELECT strftime('%Y-%m-%d', t.date, 'weekday 6') AS week_start, w.name AS worker, 
               COALESCE(SUM(t.hours), 0) AS hours
        FROM timesheets t
        JOIN workers w ON w.id = t.worker_id
        WHERE substr(t.date, 1, 7) = ?
        GROUP BY week_start, w.name
        ORDER BY week_start, w.name
    """, (month,)).fetchall()

    # Latest non-empty remark per worker in this month
    rmk_rows = conn.execute("""
        SELECT w.name AS worker, t.note, t.date
        FROM timesheets t
        JOIN workers w ON w.id=t.worker_id
        WHERE substr(t.date, 1, 7) = ? AND t.note IS NOT NULL AND t.note <> ''
        ORDER BY t.date
    """, (month,)).fetchall()

    conn.close()

    # Build full list of weeks in the month (weeks start from Saturday)
    y, m = map(int, month.split("-"))
    last = calendar.monthrange(y, m)[1]
    grid_dates = [f"{y:04d}-{m:02d}-{d:02d}" for d in range(1, last + 1)]

    # Workers to show (active + any with entries)
    grid_workers = sorted({w["name"] for w in workers} | {r["worker"] for r in week_rows})

    # Fill matrix with zeros first, then set sums (grouping by week start date)
    grid_data = {w: {d: 0.0 for d in grid_dates} for w in grid_workers}
    for r in week_rows:
        grid_data[r["worker"]][r["week_start"]] = float(r["hours"] or 0.0)

    # Calculate weekly totals (sum hours worked per week for each worker)
    grid_totals = {w: sum(grid_data[w][d] for d in grid_dates) for w in grid_workers}

    # Collect latest remarks for each worker
    grid_remarks = {w: "" for w in grid_workers}
    for r in rmk_rows:
        grid_remarks[r["worker"]] = r["note"]

    return render_template(
        "timesheets.html",
        date=the_day,
        workers=workers,
        rows=rows,
        grid_dates=grid_dates,
        grid_workers=grid_workers,
        grid_data=grid_data,
        grid_totals=grid_totals,
        grid_remarks=grid_remarks
    )

@app.route("/timesheets/approve/<int:tid>")
@login_required
@roles_required('md','admin')
def approve_timesheet(tid):
    conn = db()
    conn.execute("UPDATE timesheets SET status='approved', approved_by=?, approved_at=? WHERE id=?",
                 (current_user.id, datetime.now().isoformat(timespec='seconds'), tid))
    conn.commit(); conn.close()
    flash("Timesheet approved.", "success")
    return redirect(request.referrer or url_for("timesheets_page"))

@app.route("/timesheets/reset/<int:tid>")
@login_required
@roles_required('md','admin')
def reset_timesheet(tid):
    conn = db()
    conn.execute(
        "UPDATE timesheets SET status='pending', approved_by=NULL, approved_at=NULL WHERE id=?",
        (tid,),
    )
    conn.commit(); conn.close()
    flash("Timesheet reset to pending.", "info")
    return redirect(request.referrer or url_for("timesheets_page"))

# Reject = DELETE (remove from system)
@app.route("/timesheets/delete/<int:tid>")
@login_required
@roles_required('md','admin')
def delete_timesheet(tid):
    conn = db()
    conn.execute("DELETE FROM timesheets WHERE id=?", (tid,))
    conn.commit(); conn.close()
    flash("Timesheet deleted.", "warning")
    return redirect(request.referrer or url_for("timesheets_page"))

# Back-compat: if any template still calls /timesheets/reject/<id>, treat as delete.
@app.route("/timesheets/reject/<int:tid>")
@login_required
@roles_required('md','admin')
def reject_timesheet(tid):
    return delete_timesheet(tid)

# ----------------------- Users -----------------------

@app.route("/users", methods=["GET","POST"])
@login_required
@roles_required('md','admin')
def users_page():
    conn = db()
    if request.method == "POST":
        u = request.form["username"].strip()
        r = request.form["role"]
        p = generate_password_hash(request.form["password"])
        conn.execute("INSERT INTO users(username,password,role) VALUES(?,?,?)",(u,p,r))
        conn.commit(); flash("User created.", "success")
        conn.close()
        return redirect(url_for("users_page"))
    users = conn.execute("SELECT id,username,role FROM users ORDER BY role,username").fetchall()
    conn.close()
    return render_template("users.html", users=users)

@app.route("/users/delete/<int:uid>")
@login_required
@roles_required('md','admin')   # harden route: only MD/Admin
def delete_user(uid):
    if current_user.role == "manager":
        flash("Not allowed.", "danger"); return redirect(url_for("users_page"))
    conn = db()
    target = conn.execute("SELECT role FROM users WHERE id=?", (uid,)).fetchone()
    if not target:
        conn.close(); return redirect(url_for("users_page"))
    if current_user.role == "md" and target["role"] != "manager":
        conn.close(); flash("MD can only delete managers.", "warning"); return redirect(url_for("users_page"))
    conn.execute("DELETE FROM users WHERE id=?", (uid,))
    conn.commit(); conn.close()
    flash("Deleted.", "warning"); return redirect(url_for("users_page"))

# ----------------------- Change Password -----------------------

@app.route("/change_password", methods=["GET","POST"])
@login_required
def change_password():
    if request.method == "POST":
        newp = request.form["password"]
        conn = db(); conn.execute("UPDATE users SET password=? WHERE id=?",
                                  (generate_password_hash(newp), current_user.id))
        conn.commit(); conn.close()
        flash("Password updated.", "success")
        return redirect(url_for("welcome"))
    return render_template("change_password.html")

# ----------------------- Summary -----------------------

@app.route("/summary")
@login_required
def summary():
    start = request.args.get("start")
    end   = request.args.get("end")
    conn = db()
    if not start or not end:
        m = request.args.get("month") or _ym()
        y, mo = map(int, m.split("-"))
        last = calendar.monthrange(y, mo)[1]
        start = f"{y:04d}-{mo:02d}-01"; end = f"{y:04d}-{mo:02d}-{last:02d}"
    r = conn.execute("""
      SELECT
        COALESCE(SUM(total_cost + fixed_cost),0) AS expenses,
        COALESCE(SUM(cash_receive + add_amount - less_amount),0) AS revenue
      FROM daily_cash
      WHERE date BETWEEN ? AND ?
    """, (start, end)).fetchone()
    conn.close()
    return render_template("summary.html", start=start, end=end, expenses=r["expenses"], revenue=r["revenue"])

# ----------------------- Exports -----------------------

def _xlsx_send(wb, name):
    fp = os.path.join(EXPORTS, name); wb.save(fp)
    return send_file(fp, as_attachment=True, download_name=name)

def _date_range_from_args():
    start = request.args.get("start"); end = request.args.get("end")
    if start and end: return start, end
    m = request.args.get("month") or _ym()
    y, mo = map(int, m.split("-")); last = calendar.monthrange(y, mo)[1]
    return f"{y:04d}-{mo:02d}-01", f"{y:04d}-{mo:02d}-{last:02d}"

@app.route("/export/daily.xlsx")
@login_required
def export_daily():
    start, end = _date_range_from_args()
    c = db()
    rows = c.execute("""
      SELECT * FROM daily_cash WHERE date BETWEEN ? AND ? ORDER BY date ASC, id ASC
    """, (start, end)).fetchall()
    c.close()
    wb = Workbook(); ws = wb.active; ws.title="Daily Inputs"
    ws.append(["Date","Total Kg","Amount","GL pay","Staff","Labour","Prod cost","Coal","Diesel","Electricity","Other","Total cost",
               "Capital","Machineries","Assets","Construction","Fixed cost","Grand total",
               "Cash receive","Add","Less","Net cash","Note","Status"])
    ws.row_dimensions[1].font = Font(bold=True)
    for r in rows:
        ws.append([r["date"], r["total_kg"], r["amount_tk"], r["green_leaf_bill_payment"], r["staff_salary"], r["labour_bill"],
                   r["production_cost"], r["coal"], r["diesel"], r["electricity"], r["other_exp"], r["total_cost"],
                   r["capital_cost"], r["machineries"], r["assets_purchase"], r["construction"], r["fixed_cost"], r["grand_total"],
                   r["cash_receive"], r["add_amount"], r["less_amount"], r["net_cash"], r["note"], r["status"]])
    return _xlsx_send(wb, f"Daily_Inputs_{start}_to_{end}.xlsx")

@app.route("/export/people.xlsx")
@login_required
def export_people():
    start, end = _date_range_from_args()
    c = db()
    workers = c.execute("SELECT * FROM workers ORDER BY active DESC, name").fetchall()
    staff   = c.execute("SELECT * FROM staff ORDER BY name").fetchall()
    c.close()
    wb = Workbook(); ws = wb.active; ws.title="Workers"
    ws.append(["Name","Join date","Leave date","Active","Note"]); ws.row_dimensions[1].font=Font(bold=True)
    for w in workers: ws.append([w["name"], w["join_date"], w["leave_date"] or "", "Yes" if w["active"] else "No", w["note"] or ""])
    ws2 = wb.create_sheet("Staff")
    ws2.append(["Name","Position","Salary","Join date","Leave date"]); ws2.row_dimensions[1].font=Font(bold=True)
    for s in staff: ws2.append([s["name"], s["position"], s["salary"], s["join_date"], s["leave_date"] or ""])
    return _xlsx_send(wb, f"People_{start}_to_{end}.xlsx")

@app.route("/export/timesheets_matrix.xlsx")
@login_required
def export_timesheets_matrix():
    # Supports ?start=YYYY-MM-DD&end=YYYY-MM-DD or ?month=YYYY-MM
    start, end = _date_range_from_args()

    d0 = datetime.strptime(start, "%Y-%m-%d")
    d1 = datetime.strptime(end,   "%Y-%m-%d")

    # inclusive date list
    dates = []
    cur = d0
    while cur <= d1:
        dates.append(cur.strftime("%Y-%m-%d"))
        cur += timedelta(days=1)

    c = db()
    rows = c.execute("""
      SELECT t.date, w.name AS worker, t.hours
      FROM timesheets t
      JOIN workers w ON w.id = t.worker_id
      WHERE t.date BETWEEN ? AND ?
      ORDER BY t.date, w.name
    """, (start, end)).fetchall()
    c.close()

    workers = sorted({r["worker"] for r in rows})
    mat = {w: {d: 0.0 for d in dates} for w in workers}
    for r in rows:
        mat[r["worker"]][r["date"]] += float(r["hours"] or 0)

    wb = Workbook()
    ws = wb.active
    ws.title = "Working Hours"

    header = ["Worker"] + dates + ["Total", "Remark"]
    ws.append(header)
    ws.row_dimensions[1].font = Font(bold=True)

    for w in workers:
        daily = [mat[w][d] for d in dates]
        tot = sum(daily)
        ws.append([w] + daily + [tot, ""])

    return _xlsx_send(wb, f"Working_Hours_Matrix_{start}_to_{end}.xlsx")

@app.route("/export/timesheets.xlsx")
@login_required
def export_timesheets():
    date_arg = request.args.get("date")
    c = db()
    if date_arg:
        rows = c.execute("""
          SELECT t.date, w.name, t.hours, t.status, t.note
          FROM timesheets t JOIN workers w ON w.id=t.worker_id
          WHERE t.date=? ORDER BY t.date,w.name
        """, (date_arg,)).fetchall()
        name = f"Working_Hours_{date_arg}.xlsx"
    else:
        start, end = _date_range_from_args()
        rows = c.execute("""
          SELECT t.date, w.name, t.hours, t.status, t.note
          FROM timesheets t JOIN workers w ON w.id=t.worker_id
          WHERE t.date BETWEEN ? AND ? ORDER BY t.date,w.name
        """, (start, end)).fetchall()
        name = f"Working_Hours_{start}_to_{end}.xlsx"
    c.close()
    wb = Workbook(); ws = wb.active; ws.title="Working Hours"
    ws.append(["Date","Worker","Hours","Status","Remark"]); ws.row_dimensions[1].font=Font(bold=True)
    for r in rows: ws.append([r["date"], r["name"], r["hours"], r["status"], r["note"] or ""])
    return _xlsx_send(wb, name)

@app.route("/export/summary.xlsx")
@login_required
def export_summary():
    start, end = _date_range_from_args()
    c = db()
    r = c.execute("""
      SELECT COALESCE(SUM(total_cost + fixed_cost),0), COALESCE(SUM(cash_receive + add_amount - less_amount),0)
      FROM daily_cash WHERE date BETWEEN ? AND ?
    """, (start, end)).fetchone()
    c.close()
    expenses, revenue = r[0], r[1]
    wb = Workbook(); ws = wb.active; ws.title="Summary"
    ws.append(["From", start]); ws.append(["To", end]); ws.append([])
    ws.append(["Total Expenses", expenses]); ws.append(["Total Revenue", revenue])
    return _xlsx_send(wb, f"Monthly_Summary_{start}_to_{end}.xlsx")

# ----------------------- Run -----------------------

if __name__ == "__main__":
    print("Talma Prime running on http://127.0.0.1:8060")
    app.run(host="0.0.0.0", port=8060, debug=True)
